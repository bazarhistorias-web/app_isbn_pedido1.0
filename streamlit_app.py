import io
import re
from typing import Dict, List, Tuple

import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ==========================
# Configuración básica
# ==========================
st.set_page_config(page_title="app_isbn_pedido1.0", page_icon="📚", layout="wide")
st.title("📚 app_isbn_pedido1.0 – Actualiza columna I (Pedido) usando ISBN/SKU en B + cantidades pegadas")
st.caption("Sube tu Excel .xlsx, pega código y cantidad por línea, y descarga el archivo actualizado. La app también reporta los códigos NO detectados.")

# ==========================
# Utilidades
# ==========================

def normalize_code(x) -> str:
    if x is None:
        return ""
    return re.sub(r"[^0-9A-Za-z]", "", str(x)).upper()


def parse_pasted_quantities(text: str) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        parts = re.split(r"[\t,; ]+", line, maxsplit=1)
        code = normalize_code(parts[0])
        if not code:
            continue
        qty = 1
        if len(parts) == 2 and parts[1].strip():
            m = re.search(r"-?\d+(?:[.,]\d+)?", parts[1])
            if m:
                qtxt = m.group(0).replace(",", ".")
                try:
                    qty = int(round(float(qtxt)))
                except Exception:
                    qty = 1
        if qty < 0:
            qty = 0
        mapping[code] = mapping.get(code, 0) + qty
    return {k: v for k, v in mapping.items() if v > 0}

# ==========================
# Entrada de usuario
# ==========================
col_up, col_paste = st.columns([1, 1])
with col_up:
    up = st.file_uploader("Sube tu Excel (.xlsx)", type=["xlsx"], accept_multiple_files=False)
with col_paste:
    pasted = st.text_area(
        "Pega CÓDIGO + CANTIDAD (uno por línea; separador tab/espacio/coma/punto y coma)",
        height=220,
        placeholder="9786076380574\t1\n9786075783888\t3\n7503055508840\t1",
    )

st.divider()

c1, c2, c3 = st.columns([1, 1, 1])
with c1:
    header_row = st.number_input("Fila de encabezados (por hoja)", min_value=1, value=1, step=1)
with c2:
    overwrite = st.checkbox("Sobrescribir valores existentes en columna I", value=True)
with c3:
    clear_non_matches = st.checkbox("Limpiar columna I cuando el código NO esté en la lista pegada", value=False)

st.info("Supuestos fijos: **ISBN/SKU está en columna B** y **Pedido en columna I** en TODAS las hojas del libro.")

# ==========================
# Procesamiento
# ==========================
if up and pasted:
    code_qty = parse_pasted_quantities(pasted)
    if not code_qty:
        st.warning("No se detectaron pares código+cantidad válidos en el texto pegado.")
    else:
        st.write(f"🔎 Códigos únicos detectados en el pegado: **{len(code_qty)}**")
        data = up.read()
        wb = load_workbook(io.BytesIO(data))

        encontrados: List[Tuple[str, int, str, int, str]] = []
        detectados_en_al_menos_una = set()
        total_marks = 0
        total_clears = 0

        for ws_name in wb.sheetnames:
            ws: Worksheet = wb[ws_name]
            isbn_col = 2
            pedido_col = 9

            for r in range(int(header_row) + 1, ws.max_row + 1):
                raw_code = ws.cell(row=r, column=isbn_col).value
                code_norm = normalize_code(raw_code)
                pedido_cell = ws.cell(row=r, column=pedido_col)

                if code_norm in code_qty:
                    qty = code_qty[code_norm]
                    prev = pedido_cell.value
                    if overwrite or (prev in (None, "")):
                        pedido_cell.value = qty
                        total_marks += 1
                    encontrados.append((ws_name, r, code_norm, qty, str(prev) if prev is not None else ""))
                    detectados_en_al_menos_una.add(code_norm)
                else:
                    if clear_non_matches and (pedido_cell.value not in (None, "")):
                        pedido_cell.value = None
                        total_clears += 1

        no_detectados = [code for code in code_qty.keys() if code not in detectados_en_al_menos_una]

        st.success(f"Proceso terminado. Filas actualizadas en I: {total_marks} · Celdas I limpiadas: {total_clears}")

        tab1, tab2, tab3 = st.tabs(["✅ Encontrados", "❗ No detectados", "⬇️ Descargas"])

        import pandas as pd

        with tab1:
            if encontrados:
                df_found = pd.DataFrame(encontrados, columns=["hoja", "fila", "codigo", "cantidad_aplicada", "valor_I_anterior"])
                st.dataframe(df_found, use_container_width=True, height=360)
            else:
                st.info("No hubo coincidencias con códigos del Excel.")

        with tab2:
            if no_detectados:
                df_nf = pd.DataFrame({"codigo_no_detectado": no_detectados, "cantidad": [code_qty[c] for c in no_detectados]})
                st.dataframe(df_nf, use_container_width=True, height=360)
            else:
                st.success("¡Todos los códigos pegados fueron detectados en alguna hoja!")

        with tab3:
            out_buf = io.BytesIO()
            wb.save(out_buf)
            out_buf.seek(0)
            st.download_button(
                label="⬇️ Descargar Excel actualizado (.xlsx)",
                data=out_buf,
                file_name=f"{up.name.rsplit('.',1)[0]}_ACTUALIZADO.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            if encontrados:
                df_found = pd.DataFrame(encontrados, columns=["hoja", "fila", "codigo", "cantidad_aplicada", "valor_I_anterior"])
                st.download_button(
                    label="⬇️ Descargar coincidencias (CSV)",
                    data=df_found.to_csv(index=False).encode("utf-8"),
                    file_name="coincidencias.csv",
                    mime="text/csv",
                )
            if no_detectados:
                df_nf = pd.DataFrame({"codigo_no_detectado": no_detectados, "cantidad": [code_qty[c] for c in no_detectados]})
                st.download_button(
                    label="⬇️ Descargar no detectados (CSV)",
                    data=df_nf.to_csv(index=False).encode("utf-8"),
                    file_name="no_detectados.csv",
                    mime="text/csv",
                )

st.divider()
st.markdown(
    """
**Notas**
- Columnas fijas: **B = ISBN/SKU**, **I = Pedido**.
- Si el encabezado no está en la fila 1, ajusta la "Fila de encabezados".
- Al pegar, usa `CODIGO<TAB>cantidad` o `CODIGO cantidad`. Si falta cantidad, se asume 1. Si repites códigos, las cantidades se **suman**.
- La app reporta cuáles códigos **no se detectaron** en ninguna hoja.
    """
)
